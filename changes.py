from openpyxl import Workbook, load_workbook, utils
from table2ascii import table2ascii as t2a, PresetStyle
from xls2xlsx import XLS2XLSX
from pathlib import Path
import re, ujson

class Change:
    def __init__(self, group: str, lesson: int, src: str, dst: str, teacher: str, room: str):
        self.group = group
        self.lesson = lesson
        self.room = room
        self.teacher = teacher
        self.replacing = (src, dst)
    
    def __str__(self) -> str:
        return f"{self.group}|{self.lesson}|[{self.replacing[0]}]=>[{self.replacing[1]}]|{self.room}|{self.teacher}"
    
    def as_list(self) -> list:
        return [self.group, self.lesson, self.replacing, self.teacher, self.room]

    def as_dict(self) -> dict:
        return {
            "room": self.room,
            "group": self.group,
            "lesson": self.lesson,
            "teacher": self.teacher,
            "replace": dict(zip(('from', 'to'), self.replacing))
            }
    
    @staticmethod
    def from_dict(data: dict) -> 'Change':
        return Change(data['group'], data['lesson'], data['replace']['from'], data['replace']['to'], data['teacher'], data['room'])

class Changes:
    def __init__(self):
        self.changes = list()
    
    def __str__(self) -> str:
        return "\n".join(map(str, self.changes))
    
    def append(self, elem: Change):
        if elem.group == "Группа": return
        self.changes.append(elem)
    
    def get_by_group(self, group: str) -> list:
        res = list()
        for i in self.changes:
            if i.group == group:
                res.append(i)
        return res

    def as_list(self) -> list:
        result = []
        for i in self.changes:
            result.append(i.as_list())
        return result
    
    def as_dict(self) -> dict:
        result = {}
        for i in self.changes:
            t = i.as_dict()
            if t['group'] not in result.keys():
                result[t['group']] = {"changes":[]}
            result[t['group']]['changes'].append(t)
        return result

    @staticmethod
    def from_dict(data: dict, day: str) -> 'Changes':
        c = data['changes']
        print(c)
        res = Changes()
        res.changes = []
        for i in c:
            res.changes.append(Change.from_dict(i))
        return res

def parseFile(file: str, cache: str) -> str:
    filePL = Path(file) #ch_data
    x2x = XLS2XLSX(str(filePL.parent)+'/changes.xls')
    x2x.to_xlsx(str(filePL.parent)+'/changes.xlsx')

    wb = load_workbook(str(filePL.parent)+'/changes.xlsx', True)

    ws = wb.active
    row_count    = ws.max_row    #type: ignore
    column_count = ws.max_column #type: ignore

    print(f"{column_count}x{row_count}")

    changes = Changes()
    for row in range(row_count):
        line = list()
        for column in range(column_count):
            v = ws[utils.cell.get_column_letter(column+1)+str(row)].value #type: ignore
            if v: line.append(v)
        if len(line) == column_count-2:
            change = Change(*line)
            changes.append(change)

    data = changes.as_dict()
    with open(cache+'ch_data.json', 'w+', encoding='utf-8') as f:
        ujson.dump(data, f, ensure_ascii=False)
    
    with open(cache+'ch_meta_data.json', 'w+', encoding='utf-8') as f:
        ujson.dump({"groups": list(data.keys())}, f, ensure_ascii=False)
        
    return cache

def parseCache(filename: str, group: str, day: str) -> Changes:
    with open(filename+'ch_data.json', 'r', encoding='utf-8') as f:
        ch = ujson.load(f)
    g = ch.get(group, {})
    return Changes.from_dict(g, day)

def makeTable(data: list) -> str:
    return t2a(
        header=["Group", "Lesson", "Replace", "Teacher", "Room"],
        body=data, first_col_heading=True
    )

if __name__ == '__main__':
    filename = input("Filename: ")
    print(filename)
    cf = input("Cache: ")
    print(cf)
    qGroup = input("Group: ")
    print(qGroup)
    qDay = input("Day of week: ")
    print(qDay)

    cf = parseFile(filename, cf)
    data = parseCache(cf, qGroup, qDay)
    
    # print(data.as_dict())
    print(makeTable(data.as_list()))