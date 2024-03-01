from openpyxl import Workbook, load_workbook, utils
from table2ascii import table2ascii as t2a, PresetStyle
import re, ujson
#5min
class Lesson:
    def __init__(self, lesson: str, teacher: str, room: str):
        self.lesson = lesson
        self.teacher = teacher
        self.room = room
    
    def __bool__(self) -> bool:
        return bool(self.lesson and self.room and self.teacher)
    
    def __str__(self) -> str:
        if not self: return "Нету"
        return f"{self.lesson} [{self.room}] ({self.teacher})"
    
    def as_list(self) -> list:
        return [self.lesson or "Нету", self.teacher or "-", self.room or "-"]

    def as_dict(self) -> dict:
        return {"lesson": self.lesson or "Нету", "teacher": self.teacher or "-", "room": self.room or "-"}

    @staticmethod
    def from_dict(data: dict) -> 'Lesson':
        return Lesson(data['lesson'], data['teacher'], data['room'])

class Lessons:
    def __init__(self, day):
        self.day = day
        self.lessons = list()
    
    def __str__(self) -> str:
        return "\n".join(map(lambda x: f"{x[0]} - {str(x[1])}", self.lessons))
    
    def append(self, elem: Lesson):
        self.lessons.append((len(self.lessons)+1, elem))
        
    def as_list(self) -> list:
        res = []
        for les in self.lessons:
            res.append([les[0], *les[1].as_list()])
        return res

    def as_dict(self) -> dict:
        res = {"lessons":[]}
        for les in self.lessons:
            res["lessons"].append({"n": les[0]} | les[1].as_dict())
        return res

    @staticmethod
    def from_dict(data: dict, day: str) -> 'Lessons':
        res = Lessons(day)
        c = data.get('lessons', {})
        res.lessons = []
        for i in c:
            try: res.lessons.append((i['n'], Lesson.from_dict(i)))
            except KeyError: pass
        return res

class TimeTable:
    def __init__(self):
        self.content = {}
    
    def add(self, lessons: Lessons, day: str, group: str):
        if group not in self.content.keys(): self.content[group] = {}
        if day not in self.content[group].keys(): self.content[group][day] = lessons
    
    def as_dict(self) -> dict:
        res = {"groups": {}}
        for g, c in self.content.items():
            t = {k: v.as_dict() for k, v in c.items()}
            res["groups"][g] = t
        return res

def isGroup(s: str) -> bool:
    return bool(re.compile(r"([а-я]{1,2}\s\d{1}\.\d{1,2}-\d{2}\b( \(\d{2}\))?)").match(s))

def isDay(s:str) -> bool:
    return s.lower() in ("понедельник","вторник","среда","четверг","пятница","суббота","воскресенье")

def parseFile(file: str, cache: str) -> str:
    wb = load_workbook(file, True)

    ws = wb.active
    row_count    = ws.max_row    #type: ignore
    column_count = ws.max_column #type: ignore
    
    #INFO
    print(f"Table size: {column_count}x{row_count}")
    days = {}
    groups = {}
    for row in range(row_count):
        for column in range(column_count):
            v = ws[utils.cell.get_column_letter(column+1)+str(row)].value #type: ignore
            if v:
                v = str(v)
                if isDay(v.strip().lower()): days[v] = (column, row)
                if isGroup(v.strip().lower()): groups[v] = (column, row)
    
    tt = TimeTable()
    for g, (gc, _) in groups.items():
        for d, (_, dr) in days.items():
            g = g.strip()
            d = d.strip()
            lessons_day = Lessons(d)
            for i in range(0, 11, 2):
                lessons_day.append(
                    Lesson(
                        ws[utils.cell.get_column_letter(gc+2)+str(dr+i  )].value, #type: ignore
                        ws[utils.cell.get_column_letter(gc+2)+str(dr+i+1)].value, #type: ignore
                        ws[utils.cell.get_column_letter(gc+1)+str(dr+i  )].value  #type: ignore
                    )
                )
            tt.add(lessons_day, d, g)
            print(d, g, 'parsed')
    print("Parsing finished")
    
    with open(cache+'tt_data.json', 'w+', encoding='utf-8') as f:
        ujson.dump(tt.as_dict(), f, ensure_ascii=False)
    
    with open(cache+'tt_meta_data.json', 'w+', encoding='utf-8') as f:
        ujson.dump({"days": list(map(str.strip, days.keys())), "groups": list(groups.keys())}, f, ensure_ascii=False)
    
    return cache

def parseCache(filename: str, group: str, day: str) -> Lessons:
    with open(filename+'tt_data.json', 'r', encoding='utf-8') as f:
        tt = ujson.load(f)['groups']
    g = tt.get(group, {})
    d = g.get(day, {})
    return Lessons.from_dict(d, day)

def makeTable(data: list) -> str:
    return t2a(
        header=["№", "Lesson", "Teacher", "Room"],
        body=data, first_col_heading=True
    )

if __name__ == '__main__':
    filename = input("Filename: ")
    print(filename)
    cf = input("Cache: ")
    print(cf)
    qGroup = input("Group: ")
    print(qGroup)
    qDay = input("Day of week: ")
    print(qDay)

    cf = parseFile(filename, cf)
    data = parseCache(cf, qGroup, qDay)
    
    # print(data.as_dict())
    print(makeTable(data.as_list()))